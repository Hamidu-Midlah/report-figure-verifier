"""Deterministic tests for exact source-value cell evidence.

Run with:  python -m tests.test_source_evidence   (from the repo root)
No network or API access is required.
"""

import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

from verifier.agent import _apply_log_finding
from verifier.tools import (
    SourceWorkbook,
    FindingsLog,
    _excel_cell_ref,
    compare_source_value,
    resolve_source_sheet,
    sheets_referenced,
    with_cell_prefix,
)


def _make_workbook() -> str:
    """Adoption/Customers plus a spaced sheet name, mirroring the sample shape."""
    wb = openpyxl.Workbook()
    adoption = wb.active
    adoption.title = "Adoption"
    adoption.append(["Category", "2024", "2025", "2026"])
    adoption.append(["Firms scaling AI (%)", 32, 36, 39])

    customers = wb.create_sheet("Customers")
    customers.append(["Segment", "Count"])
    customers.append(["Enterprise", 1200])

    regional = wb.create_sheet("Regional Sales")
    regional.append(["Area", "Value"])
    regional.append(["North region", 42])

    path = os.path.join(tempfile.mkdtemp(), "test_book.xlsx")
    wb.save(path)
    return path


WB_PATH = _make_workbook()


def _finding(**kw) -> dict:
    base = {"verdict": "match", "sheet": "", "source_location": ""}
    base.update(kw)
    return base


def _flow(wb, log, query, cell, reported, *, tolerance=0.5, verdict="mismatch",
          claim_id="C1", source_location="Adoption sheet, Firms scaling AI (%), 2026 column",
          log_source_value=None, log_source_cell=None, log_sheet=None):
    """Drive find -> select -> compare -> log; return each step's result."""
    hits = wb.find_value(query)
    match = hits[0]
    sel = wb.select_source_cell(match["match_id"], cell)
    svid = sel["source_value_id"]
    cmp = compare_source_value(wb, reported, svid, tolerance)
    args = {
        "claim_id": claim_id, "sentence": "s", "reported_value": reported,
        "source_value": str(cmp["source_value"]) if log_source_value is None else log_source_value,
        "source_location": source_location, "verdict": verdict,
        "source_value_id": svid,
    }
    if log_source_cell is not None:
        args["source_cell"] = log_source_cell
    if log_sheet is not None:
        args["sheet"] = log_sheet
    res = _apply_log_finding(args, wb, log)
    return match, sel, cmp, res


# --- 1: label cell and numeric source cell are distinguished -----------------

def test_label_and_source_cell_distinguished():
    wb = SourceWorkbook(WB_PATH)
    hits = wb.find_value("scaling")
    match = hits[0]
    assert match["label_cell"] == "Adoption!A2"
    # the row exposes every cell with value + header
    d2 = next(c for c in match["cells"] if c["cell"] == "Adoption!D2")
    assert d2["value"] == 39 and d2["header"] == "2026"
    sel = wb.select_source_cell(match["match_id"], "Adoption!D2")
    assert sel["cell"] == "Adoption!D2"
    assert sel["label_cell"] == "Adoption!A2"
    assert sel["cell"] != sel["label_cell"]


# --- 2: Adoption label at A2, 2026 value at D2 -> finding records D2 ----------

def test_finding_records_value_cell_not_label_cell():
    wb = SourceWorkbook(WB_PATH)
    log = FindingsLog()
    match, sel, cmp, res = _flow(wb, log, "scaling", "Adoption!D2", "36")
    assert res.get("logged"), res
    f = log.to_dicts()[0]
    assert f["source_cell"] == "Adoption!D2"
    assert f["label_cell"] == "Adoption!A2"
    assert f["sheet"] == "Adoption"
    assert f["source_mapping_status"] == "structured"
    assert f["source_location"].startswith("Adoption!D2;")
    assert "Adoption!A2" not in f["source_location"]
    # inspectable evidence chain: match_id -> source_value_id -> comparison_id -> finding
    assert f["source_value_id"] == sel["source_value_id"]
    assert f["comparison_id"] == cmp["comparison_id"]
    assert f["source_value_id"] and f["comparison_id"]


# --- 3: exact source value comes from the workbook registry ------------------

def test_source_value_comes_from_registry():
    wb = SourceWorkbook(WB_PATH)
    log = FindingsLog()
    _m, _s, cmp, _res = _flow(wb, log, "scaling", "Adoption!D2", "36")
    assert cmp["source_value"] == 39
    assert cmp["source_cell"] == "Adoption!D2"
    assert cmp["column_header"] == "2026"
    assert cmp["verdict"] == "mismatch"
    assert log.to_dicts()[0]["source_value"] == "39"


# --- 4: a fabricated source_value_id is rejected -----------------------------

def test_fabricated_source_value_id_rejected():
    wb = SourceWorkbook(WB_PATH)
    assert "error" in compare_source_value(wb, "36", "srcval_9999", 0.5)
    log = FindingsLog()
    res = _apply_log_finding({
        "claim_id": "C1", "sentence": "s", "reported_value": "36",
        "source_value": "39", "source_location": "Adoption sheet",
        "verdict": "mismatch", "source_value_id": "srcval_9999",
    }, wb, log)
    assert "error" in res and len(log.findings) == 0


# --- 5: a cell outside the registered match is rejected ----------------------

def test_cell_outside_match_rejected():
    wb = SourceWorkbook(WB_PATH)
    match = wb.find_value("scaling")[0]
    res = wb.select_source_cell(match["match_id"], "Adoption!Z9")
    assert "error" in res


# --- 5b: source_value_id must have passed through compare_values -------------

def test_source_value_id_must_be_compared():
    wb = SourceWorkbook(WB_PATH)
    match = wb.find_value("scaling")[0]
    sel = wb.select_source_cell(match["match_id"], "Adoption!D2")
    log = FindingsLog()
    res = _apply_log_finding({
        "claim_id": "C1", "sentence": "s", "reported_value": "36",
        "source_value": "39", "source_location": "Adoption sheet",
        "verdict": "mismatch", "source_value_id": sel["source_value_id"],
    }, wb, log)  # never compared
    assert "error" in res and len(log.findings) == 0


# --- 6: log_finding cannot substitute a different cell or value --------------

def test_log_cannot_substitute_cell_or_value():
    wb = SourceWorkbook(WB_PATH)
    log = FindingsLog()
    _m, _s, _c, res_val = _flow(wb, log, "scaling", "Adoption!D2", "36",
                                log_source_value="99")
    assert "error" in res_val, "conflicting source value must be rejected"

    log2 = FindingsLog()
    _m, _s, _c, res_cell = _flow(wb, log2, "scaling", "Adoption!D2", "36",
                                 log_source_cell="Adoption!B2")
    assert "error" in res_cell, "conflicting source cell must be rejected"

    log3 = FindingsLog()
    _m, _s, _c, res_sheet = _flow(wb, log3, "scaling", "Adoption!D2", "36",
                                  log_sheet="Customers")
    assert "error" in res_sheet, "conflicting sheet must be rejected"


# --- 7: sheet names with spaces / apostrophes remain valid -------------------

def test_special_sheet_names_valid():
    assert _excel_cell_ref("Regional Sales", 3, 4) == "'Regional Sales'!D3"
    assert _excel_cell_ref("O'Brien", 1, 1) == "'O''Brien'!A1"
    wb = SourceWorkbook(WB_PATH)
    log = FindingsLog()
    match = wb.find_value("North region")[0]
    assert match["sheet"] == "Regional Sales"
    value_cell = next(c["cell"] for c in match["cells"] if c["value"] == 42)
    assert value_cell.startswith("'Regional Sales'!")
    _m, _s, _c, res = _flow(wb, log, "North region", value_cell, "40",
                            source_location="Regional Sales sheet, North region")
    assert res.get("logged"), res
    f = log.to_dicts()[0]
    assert f["sheet"] == "Regional Sales" and f["source_cell"] == value_cell


# --- 8: existing sheet counting still works ----------------------------------

def test_sheet_counting_still_works():
    findings = [
        _finding(sheet="Adoption"),
        _finding(sheet="Adoption"),
        _finding(sheet="Customers"),
    ]
    assert sheets_referenced(findings) == ["Adoption", "Customers"]
    assert len(sheets_referenced(findings)) == 2


# --- 9: unverifiable findings are excluded and carry no source cell ----------

def test_unverifiable_excluded():
    wb = SourceWorkbook(WB_PATH)
    log = FindingsLog()
    res = _apply_log_finding({
        "claim_id": "C1", "sentence": "external forecast",
        "reported_value": "200 billion", "source_value": "",
        "source_location": "Not present in spreadsheet", "verdict": "unverifiable",
    }, wb, log)
    assert res.get("logged"), res
    f = log.to_dicts()[0]
    assert f["sheet"] == "" and f["source_cell"] == "" and f["label_cell"] == ""
    assert f["source_value_id"] == "" and f["comparison_id"] == ""
    assert resolve_source_sheet(f) == ("", "n/a")
    assert sheets_referenced([f]) == []


# --- 10: cell prefixes are not duplicated ------------------------------------

def test_cell_prefix_not_duplicated():
    already = "Adoption!D2; Adoption sheet, 2026 column"
    assert with_cell_prefix("Adoption!D2", already) == already
    prefixed = with_cell_prefix("Adoption!D2", "Adoption sheet, 2026 column")
    assert prefixed == "Adoption!D2; Adoption sheet, 2026 column"
    assert with_cell_prefix("Adoption!D2", prefixed) == prefixed
    # a different leading cell is replaced, not stacked
    assert with_cell_prefix("Adoption!D2", "Adoption!A2; Adoption sheet") == \
        "Adoption!D2; Adoption sheet"


def main() -> int:
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    failed = 0
    for t in tests:
        try:
            t()
            print(f"PASS  {t.__name__}")
        except Exception as exc:  # noqa: BLE001
            failed += 1
            print(f"FAIL  {t.__name__}: {type(exc).__name__}: {exc}")
    print(f"\n{len(tests) - failed}/{len(tests)} tests passed")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
