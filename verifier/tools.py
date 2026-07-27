"""Deterministic tools exposed to the LLM agent via function calling.

Design principle: the LLM never does arithmetic or reads cells "from memory".
Every number it verifies must come through one of these grounded tools, and
every comparison is computed in Python, not by the model.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field, asdict

import openpyxl
from openpyxl.utils import get_column_letter


def _excel_cell_ref(sheet: str, row: int, col: int) -> str:
    """Return a sheet-qualified A1 reference, e.g. Sales!D3 or 'Regional Sales'!D3.

    Sheet names that contain anything other than letters, digits, or underscores
    are wrapped in single quotes (with any internal quote doubled), matching how
    Excel itself qualifies such references.
    """
    a1 = f"{get_column_letter(col)}{row}"
    sheet = sheet or ""
    if re.fullmatch(r"[A-Za-z0-9_]+", sheet):
        return f"{sheet}!{a1}"
    return f"'{sheet.replace(chr(39), chr(39) * 2)}'!{a1}"


def sanitize(text):
    """Replace em dashes in text with safe punctuation (this project bans them)."""
    if not isinstance(text, str):
        return text
    em = chr(8212)  # em dash (U+2014), built from its code point to keep this file clean
    return text.replace(f" {em} ", ": ").replace(em, "-")


# ---------------------------------------------------------------------------
# Spreadsheet access
# ---------------------------------------------------------------------------

class SourceWorkbook:
    """Read-only wrapper around the source-of-truth spreadsheet."""

    def __init__(self, path: str):
        self._wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        # Per-run evidence registries. Findings cite tool-issued ids, and the
        # exact sheet / cell / value are resolved here, never from model prose.
        self._match_counter = 0
        self._matches: dict[str, dict] = {}        # match_id -> row evidence
        self._srcval_counter = 0
        self._source_values: dict[str, dict] = {}  # source_value_id -> cell evidence
        self._comparison_counter = 0
        self._comparisons: dict[str, dict] = {}    # comparison_id -> comparison record

    def list_sheets(self) -> list[str]:
        return self._wb.sheetnames

    def _register_match(self, sheet: str, label_cell: str, row_label: str,
                        cell_list: list) -> str:
        self._match_counter += 1
        match_id = f"match_{self._match_counter:04d}"
        self._matches[match_id] = {
            "sheet": sheet,
            "label_cell": label_cell,
            "row_label": row_label,
            "cells": {c["cell"]: c for c in cell_list},
        }
        return match_id

    def select_source_cell(self, match_id: str, cell: str) -> dict:
        """Register the exact value cell chosen from a match; return its evidence.

        Rejects a cell that was not part of the named match. The returned
        source_value_id is what compare_values and log_finding must reference.
        """
        match = self._matches.get(match_id)
        if match is None:
            return {"error": f"Unknown match_id '{match_id}'. "
                             "Use a match_id from find_in_spreadsheet."}
        info = match["cells"].get(cell)
        if info is None:
            return {"error": f"Cell '{cell}' was not returned for match_id "
                             f"'{match_id}'. Choose a cell from that match's cells."}
        self._srcval_counter += 1
        source_value_id = f"srcval_{self._srcval_counter:04d}"
        self._source_values[source_value_id] = {
            "source_value_id": source_value_id,
            "sheet": match["sheet"],
            "cell": cell,
            "value": info["value"],
            "header": info["header"],
            "row_label": match["row_label"],
            "label_cell": match["label_cell"],
        }
        return dict(self._source_values[source_value_id])

    def resolve_source_value(self, source_value_id: str):
        """Return the registered evidence for a source_value_id, else None."""
        return self._source_values.get(source_value_id)

    def register_comparison(self, record: dict) -> str:
        """Record one compare_values call; link it back to its source value."""
        self._comparison_counter += 1
        comparison_id = f"cmp_{self._comparison_counter:04d}"
        record = dict(record)
        record["comparison_id"] = comparison_id
        self._comparisons[comparison_id] = record
        source = self._source_values.get(record.get("source_value_id"))
        if source is not None:
            source["comparison_id"] = comparison_id
        return comparison_id

    def resolve_comparison(self, comparison_id: str):
        """Return the registered comparison record, else None."""
        return self._comparisons.get(comparison_id)

    def read_sheet(self, sheet_name: str, max_rows: int = 200) -> list[list]:
        """Return the sheet as a list of rows (truncated for context safety)."""
        if sheet_name not in self._wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {self._wb.sheetnames}")
        ws = self._wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append([("" if c is None else c) for c in row])
        return rows

    def find_value(self, query: str) -> list[dict]:
        """Search all sheets for cells whose text matches `query` (case-insensitive).

        Each match carries a `match_id`, the `sheet`, the `label_cell` that was
        matched, and a `cells` list exposing every cell in that row with its
        exact `cell` reference, `value`, and column `header`. Pick the numeric
        value cell with select_source_cell; the sheet, cell, and value are then
        resolved deterministically rather than parsed from prose.
        """
        query_l = query.lower()
        raw = []
        for name in self._wb.sheetnames:
            ws = self._wb[name]
            header_row = None
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                cells = [("" if c is None else c) for c in row]
                if r_idx == 1:
                    header_row = cells
                for c_idx, cell in enumerate(row, start=1):
                    if cell is not None and query_l in str(cell).lower():
                        raw.append((name, r_idx, c_idx, header_row, cells))

        hits = []
        for name, r_idx, c_idx, header_row, row_cells in raw[:20]:
            label_ref = _excel_cell_ref(name, r_idx, c_idx)
            label_value = row_cells[c_idx - 1] if c_idx - 1 < len(row_cells) else ""
            cell_list = []
            for j, value in enumerate(row_cells, start=1):
                header = header_row[j - 1] if header_row and j - 1 < len(header_row) else ""
                cell_list.append({
                    "sheet": name,
                    "cell": _excel_cell_ref(name, r_idx, j),
                    "row": r_idx,
                    "column": get_column_letter(j),
                    "value": value,
                    "header": header,
                })
            match_id = self._register_match(name, label_ref, str(label_value), cell_list)
            hits.append({
                "match_id": match_id,
                "sheet": name,
                "label_cell": label_ref,
                "row": r_idx,
                "header_row": header_row,
                "cells": cell_list,
            })
        return hits  # capped to 20 results above


# ---------------------------------------------------------------------------
# Claim extraction from the report text
# ---------------------------------------------------------------------------

# Matches e.g. "39%", "39.5 %", "USD 4.2 billion", "4,200", "0.7pp"
# Known issue (deferred to a later numeric-normalisation change): the unit
# alternation matches "m" before "million" (and "k"/"bn" similarly), so
# "95 million" is captured as the figure "95 m". This is format-agnostic and left
# unchanged here on purpose; fix it when unit normalisation is added.
_NUMBER_PATTERN = re.compile(
    r"(?P<num>\d{1,3}(?:,\d{3})+|\d+(?:\.\d+)?)\s*(?P<unit>%|pp|percentage points?|bn|billion|m|million|k|thousand)?",
    re.IGNORECASE,
)

# Ingestion inserts "[Page N]" boundaries into PDF text for the reviewer preview.
# They are page markers, not figures, so they are ignored before claim scanning
# (the page number must never be mistaken for a numeric claim).
_PAGE_MARKER = re.compile(r"\[\s*page\s+\d+\s*\]", re.IGNORECASE)


def _looks_like_year(num: str, unit: str | None) -> bool:
    """A bare 4-digit number like 2024/2026 scopes a claim; it is not a figure."""
    return (
        not unit
        and num.isdigit()
        and "," not in num
        and 1900 <= int(num) <= 2099
    )


def extract_numeric_claims(report_text: str) -> list[dict]:
    """Pre-scan the report for numeric claims, one claim per figure.

    This is a deterministic first pass; the agent then decides which claims
    are verifiable against the spreadsheet and which are out of scope
    (e.g. sample sizes quoted from third parties).

    Reports are commonly hard-wrapped, so a claim's number and its label can
    sit on different physical lines. We unwrap each paragraph before splitting
    into sentences, so every figure stays attached to the sentence that gives
    it meaning. Each numeric figure becomes its own claim (carrying the full
    sentence) so multi-figure sentences are still verified figure by figure.
    Bare years are treated as scope, not as figures to verify.
    """
    claims = []
    claim_id = 0
    # Drop page-boundary markers so a page number never becomes a numeric claim.
    report_text = _PAGE_MARKER.sub(" ", report_text)
    # Paragraphs are separated by blank lines; unwrap hard line breaks inside one.
    for paragraph in re.split(r"\n\s*\n", report_text):
        paragraph = re.sub(r"\s+", " ", paragraph.replace("\n", " ")).strip()
        if not paragraph:
            continue
        for sent in re.split(r"(?<=[.!?])\s+", paragraph):
            sent = sent.strip()
            if not sent:
                continue
            for m in _NUMBER_PATTERN.finditer(sent):
                num, unit = m.group("num"), m.group("unit")
                if num is None or _looks_like_year(num, unit):
                    continue
                claim_id += 1
                claims.append({
                    "claim_id": f"C{claim_id:03d}",
                    "sentence": sent,
                    "figure": m.group(0).strip(),
                })
    return claims


# ---------------------------------------------------------------------------
# Comparison: done in Python, never by the model
# ---------------------------------------------------------------------------

def _to_float(value) -> float | None:
    if value is None:
        return None
    s = str(value).replace(",", "").replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def compare_values(reported, source, tolerance: float = 0.0) -> dict:
    """Compare a reported figure against a source figure.

    tolerance is an absolute allowance (e.g. 0.5 to permit rounding to the
    nearest whole percentage point). Verdicts:
      - match           : |reported - source| <= tolerance
      - rounding_diff   : within 2.0 absolute but outside tolerance
      - mismatch        : everything else
      - not_comparable  : either side is non-numeric
    """
    r, s = _to_float(reported), _to_float(source)
    if r is None or s is None:
        return {"verdict": "not_comparable", "reported": reported, "source": source}
    diff = abs(r - s)
    if diff <= tolerance:
        verdict = "match"
    elif diff <= 2.0:
        verdict = "rounding_diff"
    else:
        verdict = "mismatch"
    return {
        "verdict": verdict,
        "reported": r,
        "source": s,
        "abs_diff": round(diff, 4),
        "tolerance": tolerance,
    }


def values_equal(a, b) -> bool:
    """Numeric equality when both parse as numbers, else exact string match."""
    fa, fb = _to_float(a), _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) < 1e-9
    return str(a).strip() == str(b).strip()


def compare_source_value(workbook, reported_value, source_value_id, tolerance=0.0) -> dict:
    """Compare a reported figure against a registered source cell's value.

    The source number is resolved from source_value_id inside Python; the model
    never supplies or retypes it. The result carries the exact source cell,
    sheet, row label, and column header alongside the verdict.
    """
    info = workbook.resolve_source_value(source_value_id)
    if info is None:
        return {"error": f"Unknown source_value_id '{source_value_id}'. "
                         "Call select_source_cell first."}
    result = compare_values(reported_value, info["value"], tolerance)
    comparison_id = workbook.register_comparison({
        "source_value_id": source_value_id,
        "reported_value": reported_value,
        "tolerance": tolerance,
        "verdict": result["verdict"],
        "source_value": info["value"],
        "source_cell": info["cell"],
        "sheet": info["sheet"],
        "label_cell": info["label_cell"],
        "row_label": info["row_label"],
        "column_header": info["header"],
        "difference": result.get("abs_diff"),
    })
    result.update({
        "comparison_id": comparison_id,
        "source_value": info["value"],
        "source_cell": info["cell"],
        "sheet": info["sheet"],
        "row_label": info["row_label"],
        "column_header": info["header"],
        "difference": result.get("abs_diff"),
    })
    return result


# ---------------------------------------------------------------------------
# Source evidence and display helpers (deterministic, no free-text parsing
# during normal operation)
# ---------------------------------------------------------------------------

# Legacy fallback only: recover "X sheet" from a human-readable source_location
# when a finding predates structured evidence.
_SHEET_TEXT_RE = re.compile(r"([A-Za-z0-9'’ _-]+?)\s+sheet\b", re.IGNORECASE)

# A leading sheet-qualified A1 reference, e.g. "Sales!D3" or "'Regional Sales'!D3".
_LEADING_CELL = re.compile(r"^\s*(?:'[^']+'|[A-Za-z0-9_]+)![A-Z]{1,3}\d+\s*;?\s*")

# Scale units the spreadsheet may imply (k / m / bn ...).
_VALUE_UNIT = re.compile(r"\d\s*(k|m|bn|billion|million|thousand)\b", re.IGNORECASE)
_CONTEXT_UNIT = re.compile(r"\b(k|m|bn|billion|million|thousand)\b", re.IGNORECASE)
_PLAIN_NUMBER = re.compile(r"^[\d,]*\.?\d+$")


def _parse_sheet_text(location: str) -> str:
    m = _SHEET_TEXT_RE.search(location or "")
    return m.group(1).strip() if m else ""


def resolve_source_sheet(finding: dict):
    """Deterministically resolve (sheet, source_mapping_status) for a finding.

    Structured `sheet` wins. Only when it is absent (a legacy finding) do we
    fall back to parsing "X sheet" from the human-readable source_location.
    """
    if finding.get("verdict") == "unverifiable":
        return "", "n/a"
    sheet = (finding.get("sheet") or "").strip()
    if sheet:
        return sheet, "structured"
    parsed = _parse_sheet_text(finding.get("source_location", ""))
    if parsed:
        return parsed, "fallback"
    return "", "under_specified"


def sheets_referenced(findings) -> list:
    """Unique source sheets cited by verifiable findings, counted once each."""
    sheets = set()
    for f in findings:
        sheet, _status = resolve_source_sheet(f)
        if sheet:
            sheets.add(sheet)
    return sorted(sheets)


def with_cell_prefix(cell: str, location: str) -> str:
    """Prepend the authoritative cell to source_location, never duplicating it."""
    location = (location or "").strip()
    if not cell:
        return location
    if location.startswith(cell):
        return location
    stripped = _LEADING_CELL.sub("", location).strip()
    return f"{cell}; {stripped}" if stripped else cell


def _unit_of(value) -> str:
    m = _VALUE_UNIT.search(str(value or ""))
    return m.group(1).lower() if m else ""


def display_source_value(finding: dict) -> str:
    """Source value for display, preserving an implied scale unit (k/m/bn)."""
    src = str(finding.get("source_value", "")).strip()
    if _unit_of(src):
        return src
    unit = _unit_of(finding.get("reported_value", ""))
    if not unit:
        m = _CONTEXT_UNIT.search(finding.get("source_location", "") or "")
        unit = m.group(1).lower() if m else ""
    if unit and _PLAIN_NUMBER.match(src):
        return f"{src}{unit}"
    return src


def mismatch_line(finding: dict) -> str:
    """A concise, deterministic one-liner for a mismatch banner entry."""
    loc = (finding.get("source_location") or "").strip()
    prefix = f"{loc}: " if loc else ""
    return sanitize(
        f"{prefix}report says {finding.get('reported_value', '')}; "
        f"spreadsheet says {display_source_value(finding)}."
    )


# ---------------------------------------------------------------------------
# Findings log: the agent's structured output channel
# ---------------------------------------------------------------------------

@dataclass
class Finding:
    claim_id: str
    sentence: str
    reported_value: str
    source_value: str
    source_location: str
    verdict: str          # match | rounding_diff | mismatch | unverifiable
    note: str = ""
    sheet: str = ""         # resolved worksheet name (empty for unverifiable)
    source_cell: str = ""   # exact numeric value cell, e.g. Adoption!D2
    label_cell: str = ""    # row-label cell used to find the row, e.g. Adoption!A2
    source_value_id: str = ""  # evidence chain: select_source_cell id (empty if unverifiable)
    comparison_id: str = ""    # evidence chain: compare_values id (empty if unverifiable)
    source_mapping_status: str = ""  # structured | fallback | under_specified | n/a


@dataclass
class FindingsLog:
    findings: list[Finding] = field(default_factory=list)

    def add(self, **kwargs) -> dict:
        f = Finding(**kwargs)
        self.findings.append(f)
        return {"logged": True, "total_findings": len(self.findings)}

    def to_dicts(self) -> list[dict]:
        return [asdict(f) for f in self.findings]

    def summary(self) -> dict:
        counts: dict[str, int] = {}
        for f in self.findings:
            counts[f.verdict] = counts.get(f.verdict, 0) + 1
        return counts
